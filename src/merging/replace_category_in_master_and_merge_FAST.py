from __future__ import annotations
import os
import re
from typing import Tuple, Set

from openpyxl import load_workbook, Workbook


# -----------------------------
# CONFIG (edit these 4 lines)
# -----------------------------
CATEGORY_TO_REPLACE = "Transportation"

MASTER_PATH = r"C:\Users\uduhe\OneDrive\Desktop\Leads_MASTER.xlsx"
CLEAN_CATEGORY_PATH = r"C:\Users\uduhe\Desktop\YP_OUTPUT\TRANSPORTATION_RECOVERED_CLEAN.xlsx"

OUTPUT_PATH = r"C:\Users\uduhe\OneDrive\Desktop\Leads_MASTER_UPDATED_TRANSPORTATION.xlsx"


# -----------------------------
# Helpers
# -----------------------------
def norm(s: str) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())

def norm_phone(s: str) -> str:
    if s is None:
        return ""
    digits = re.sub(r"\D", "", str(s))
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    return digits if len(digits) == 10 else ""

def make_key_name_addr(name: str, addr: str) -> str:
    return f"{norm(name)}|{norm(addr)}"

def safe_cell(row, idx: int):
    v = row[idx].value if idx < len(row) else None
    return "" if v is None else str(v).strip()

def build_header_map(headers) -> dict:
    return {str(h).strip(): i for i, h in enumerate(headers) if h is not None}


# -----------------------------
# Main
# -----------------------------
def main():
    print("Opening MASTER workbook (read-only)...")
    master_wb = load_workbook(MASTER_PATH, read_only=True, data_only=True)
    master_ws = master_wb.active

    # read headers
    master_rows = master_ws.iter_rows(values_only=False)
    master_header_row = next(master_rows)
    master_headers = [c.value for c in master_header_row]

    master_map = build_header_map(master_headers)
    required = ["Category", "Subcategory", "Company Name", "Address", "Phone Number"]
    missing = [c for c in required if c not in master_map]
    if missing:
        raise ValueError(f"MASTER missing required columns: {missing}")

    print("Creating NEW output workbook...")
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Leads"

    out_ws.append(required)

    # pass 1: keep non-target category rows
    kept = 0
    removed = 0

    # dedupe sets from kept
    seen_phones: Set[str] = set()
    seen_name_addr: Set[str] = set()

    total_est = master_ws.max_row or 0
    print(f"Scanning MASTER rows (total ~{total_est}) and removing '{CATEGORY_TO_REPLACE}'...")

    for idx, row in enumerate(master_rows, start=2):
        cat = safe_cell(row, master_map["Category"])
        if norm(cat) == norm(CATEGORY_TO_REPLACE):
            removed += 1
            continue

        name = safe_cell(row, master_map["Company Name"])
        addr = safe_cell(row, master_map["Address"])
        phone = safe_cell(row, master_map["Phone Number"])

        out_ws.append([cat, safe_cell(row, master_map["Subcategory"]), name, addr, phone])
        kept += 1

        pkey = norm_phone(phone)
        if pkey:
            seen_phones.add(pkey)
        else:
            seen_name_addr.add(make_key_name_addr(name, addr))

        if idx % 5000 == 0:
            print(f"  progress: {idx}/{total_est} | kept={kept} removed={removed}")

    print(f"MASTER pass done. kept={kept}, removed={removed}")
    print(f"Dedupe keys: phones={len(seen_phones)} name+addr={len(seen_name_addr)}")

    # pass 2: append clean category with dedupe
    print("Opening CLEAN category workbook...")
    clean_wb = load_workbook(CLEAN_CATEGORY_PATH, read_only=True, data_only=True)
    clean_ws = clean_wb.active

    clean_rows = clean_ws.iter_rows(values_only=True)
    clean_header = next(clean_rows)
    clean_map = build_header_map(clean_header)

    missing2 = [c for c in required if c not in clean_map]
    if missing2:
        raise ValueError(f"CLEAN file missing required columns: {missing2}")

    added = 0
    skipped = 0

    clean_total = clean_ws.max_row or 0
    print("Appending CLEAN category rows with dedupe...")

    for i, r in enumerate(clean_rows, start=2):
        cat = (r[clean_map["Category"]] or "")
        subcat = (r[clean_map["Subcategory"]] or "")
        name = (r[clean_map["Company Name"]] or "")
        addr = (r[clean_map["Address"]] or "")
        phone = (r[clean_map["Phone Number"]] or "")

        pkey = norm_phone(phone)
        if pkey:
            if pkey in seen_phones:
                skipped += 1
                continue
        else:
            nakey = make_key_name_addr(name, addr)
            if nakey in seen_name_addr:
                skipped += 1
                continue

        out_ws.append([cat, subcat, name, addr, phone])
        added += 1

        if pkey:
            seen_phones.add(pkey)
        else:
            seen_name_addr.add(make_key_name_addr(name, addr))

        if i % 1000 == 0:
            print(f"  clean progress: {i}/{clean_total} | added={added} skipped={skipped}")

    print(f"CLEAN append done. added={added}, skipped={skipped}")

    print("Saving output workbook (this can take a bit)...")
    out_wb.save(OUTPUT_PATH)
    print("Saved:", OUTPUT_PATH)


if __name__ == "__main__":
    main()
